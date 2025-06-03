def update_excel_with_results(results, file_path):
    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    ws = wb.active

    for result in results:
        tag = result["tags"][0] if result["tags"] else ""
        if not tag:
            continue

        keyword = tag

        print("Result content:", result)
        if "scenario" in result:
            print("Scenario from result:", result["scenario"])

        for row in ws.iter_rows(min_row=10):
            if row[9].value and keyword in str(row[9].value):
                for col in range(4, 9):
                    row[col].value = ""

                row[4].value = "*"

                if result["status"] == "passed":
                    row[5].value = "*"
                else:
                    row[6].value = "*"

                row[7].value = "Chromium"
                row[8].value = "Test passed" if result["status"] == "passed" else "Test failed"

                break

    wb.save(file_path)
